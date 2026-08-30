"""
upload_schemes.py
==================
SIH 2026 PS92 - Loads verified scheme + channel-partner data from the research
spreadsheet (PS92_Scheme_Research.xlsx) into Firestore, OR previews it as JSON
if you don't have Firebase credentials set up yet.

WHY THIS SCRIPT EXISTS
-----------------------
Firestore should always be a MIRROR of the verified spreadsheet, not a second
source of truth. Whenever the spreadsheet changes, re-run this script.

HOW TO USE (beginner-friendly walkthrough)
-------------------------------------------
1. First, just PREVIEW what would be uploaded, with no Firebase setup needed:

     python upload_schemes.py --dry-run

   This reads the spreadsheet and writes two files you can inspect:
     schemes_preview.json
     channel_partners_preview.json

   Check these carefully before touching real Firestore.

2. Once your Firebase project exists and you've downloaded a service account
   key (Firebase Console -> Project Settings -> Service Accounts -> Generate
   new private key), save it as serviceAccountKey.json in this same folder
   (or point to it with --credentials path/to/key.json), then run:

     pip install firebase-admin openpyxl
     python upload_schemes.py

   This uploads every VERIFIED row to two Firestore collections:
     schemes            (document ID = scheme_id, e.g. "NSFDC007")
     channel_partners   (document ID = partner_id, e.g. "CP001")

3. Rows with data_status other than "verified" (e.g. "needs_review") are
   SKIPPED automatically and printed as warnings, so a half-checked scheme
   never accidentally reaches your live app.

IMPORTANT NOTES FOR THE TEAM
------------------------------
- Some spreadsheet fields aren't clean numbers (e.g. "No income ceiling",
  "Not fixed"). For these, `income_limit` / `minimum_loan_amount` /
  `maximum_loan_amount` / `maximum_project_cost` will be stored as `null`
  in Firestore, and the ORIGINAL text is preserved in a matching
  `..._raw` field. Your eligibility engine should treat `null` as
  "no numeric limit / not applicable" rather than crashing or treating it
  as zero. Schemes needing this special handling are printed by this
  script under "NEEDS MANUAL HANDLING" when you run it.
- List-style fields (target_groups, business_types, supported_purposes,
  eligible_activities, required_documents, optional_documents) are split
  on "; " into arrays. Some list items include extra descriptive text in
  parentheses (e.g. "Trading/Business (trading capped at 10%...)") -- when
  matching a user's input against these in the eligibility engine, prefer
  a simple "is the user's word CONTAINED in this list item" check over an
  exact match.
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency. Run: pip install openpyxl --break-system-packages")

SPREADSHEET_FILE = "PS92_Scheme_Research.xlsx"
SCHEMES_SHEET = "schemes"
PARTNERS_SHEET = "channel_partners"

LIST_FIELDS = [
    "target_groups", "business_types", "supported_purposes",
    "eligible_activities", "required_documents", "optional_documents",
]

NUMERIC_FIELDS = [
    "income_limit", "minimum_loan_amount", "maximum_loan_amount",
    "maximum_project_cost", "minimum_age", "maximum_age",
]


def read_sheet_as_dicts(path, sheet_name):
    """Reads any sheet into a list of plain dicts keyed by its header row."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        sys.exit(f"Sheet '{sheet_name}' not found in {path}. Found: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c + 1).value for c in range(len(headers))]
        if all(v in (None, "") for v in values):
            continue  # skip fully blank rows
        rows.append(dict(zip(headers, values)))
    return rows


def to_number_or_none(value):
    """Tries to turn a spreadsheet cell into a clean number.
    Returns (number_or_None, raw_original_value)."""
    if isinstance(value, (int, float)):
        return value, None
    if value is None:
        return None, None
    text = str(value).strip()
    # allow plain digit strings like "300000"
    if re.fullmatch(r"\d+", text):
        return int(text), None
    return None, text


def split_list_field(value):
    """Splits a '; '-separated spreadsheet cell into a clean list of strings."""
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [item.strip() for item in text.split(";") if item.strip()]


def build_scheme_doc(row):
    doc = {
        "schemeId": row.get("scheme_id"),
        "schemeName": row.get("scheme_name"),
        "organization": row.get("organization"),
        "schemeType": row.get("scheme_type"),
        "description": row.get("description"),
        "entrepreneurshipRelevance": row.get("entrepreneurship_relevance"),
        "targetGroups": split_list_field(row.get("target_groups")),
        "genderRequirement": row.get("gender_requirement"),
        "stateRestriction": row.get("state_restriction"),
        "otherBeneficiaryConditions": row.get("other_beneficiary_conditions"),
        "beneficiaryContribution": row.get("beneficiary_contribution"),
        "interestRate": row.get("interest_rate"),
        "repaymentPeriod": row.get("repayment_period"),
        "moratoriumPeriod": row.get("moratorium_period"),
        "businessTypes": split_list_field(row.get("business_types")),
        "supportedPurposes": split_list_field(row.get("supported_purposes")),
        "eligibleActivities": split_list_field(row.get("eligible_activities")),
        "newBusinessAllowed": bool(row.get("new_business_allowed")),
        "existingBusinessAllowed": bool(row.get("existing_business_allowed")),
        "eligibilityRules": row.get("eligibility_rules"),
        "mandatoryConditions": row.get("mandatory_conditions"),
        "exclusionConditions": row.get("exclusion_conditions"),
        "requiredDocuments": split_list_field(row.get("required_documents")),
        "optionalDocuments": split_list_field(row.get("optional_documents")),
        "applicationMethod": row.get("application_method"),
        "applicationAuthority": row.get("application_authority"),
        "channelisingAgency": row.get("channelising_agency"),
        "applicationUrl": row.get("application_url"),
        "officialSourceUrl": row.get("official_source_url"),
        "sourceDocument": row.get("source_document"),
        "sourceDate": row.get("source_date"),
        "lastVerified": str(row.get("last_verified")) if row.get("last_verified") else None,
        "verificationNotes": row.get("verification_notes"),
        "dataStatus": row.get("data_status"),
    }

    # numeric fields with raw-text fallback for non-numeric cells
    manual_review_flags = []
    for field in NUMERIC_FIELDS:
        camel = re.sub(r"_([a-z])", lambda m: m.group(1).upper(), field)
        number, raw = to_number_or_none(row.get(field))
        doc[camel] = number
        if raw is not None:
            doc[camel + "Raw"] = raw
            manual_review_flags.append(field)

    return doc, manual_review_flags


def build_partner_doc(row):
    return {
        "partnerId": row.get("partner_id"),
        "partnerName": row.get("partner_name"),
        "partnerType": row.get("partner_type"),
        "state": row.get("state"),
        "district": row.get("district"),
        "schemesHandled": split_list_field(row.get("schemes_handled")),
        "address": row.get("address"),
        "contactNumber": row.get("contact_number"),
        "latitude": row.get("latitude"),
        "longitude": row.get("longitude"),
        "mockUtilizationStatus": row.get("mock_utilization_status"),
        "notes": row.get("notes"),
        "isSimulatedData": True,  # always true for this prototype -- never claim this is real
    }


def main():
    parser = argparse.ArgumentParser(description="Upload verified PS92 scheme data to Firestore.")
    parser.add_argument("--file", default=SPREADSHEET_FILE, help="Path to the research spreadsheet")
    parser.add_argument("--credentials", default="serviceAccountKey.json", help="Path to Firebase service account key")
    parser.add_argument("--dry-run", action="store_true", help="Preview as JSON instead of uploading to Firestore")
    args = parser.parse_args()

    spreadsheet_path = Path(args.file)
    if not spreadsheet_path.exists():
        sys.exit(f"Spreadsheet not found: {spreadsheet_path}")

    scheme_rows = read_sheet_as_dicts(spreadsheet_path, SCHEMES_SHEET)
    partner_rows = read_sheet_as_dicts(spreadsheet_path, PARTNERS_SHEET)

    verified_schemes = []
    skipped_schemes = []
    needs_review_fields = {}

    for row in scheme_rows:
        status = str(row.get("data_status", "")).lower()
        if status.startswith("verified"):
            doc, flags = build_scheme_doc(row)
            verified_schemes.append(doc)
            if flags:
                needs_review_fields[doc["schemeId"]] = flags
        else:
            skipped_schemes.append((row.get("scheme_id"), row.get("data_status")))

    partner_docs = [build_partner_doc(row) for row in partner_rows]

    print(f"\nRead {len(scheme_rows)} scheme rows -> {len(verified_schemes)} verified, {len(skipped_schemes)} skipped")
    for sid, status in skipped_schemes:
        print(f"  SKIPPED {sid}: data_status = '{status}' (fix in spreadsheet, then re-run)")

    if needs_review_fields:
        print("\nNEEDS MANUAL HANDLING IN YOUR ELIGIBILITY ENGINE (non-numeric limit fields):")
        for sid, fields in needs_review_fields.items():
            print(f"  {sid}: {', '.join(fields)} stored as null with a '...Raw' text field -- code around this")

    print(f"\nRead {len(partner_rows)} channel partner rows (all treated as simulated demo data)")

    if args.dry_run:
        with open("schemes_preview.json", "w", encoding="utf-8") as f:
            json.dump(verified_schemes, f, indent=2, ensure_ascii=False)
        with open("channel_partners_preview.json", "w", encoding="utf-8") as f:
            json.dump(partner_docs, f, indent=2, ensure_ascii=False)
        print("\nDRY RUN COMPLETE. Wrote schemes_preview.json and channel_partners_preview.json")
        print("Inspect these, then re-run without --dry-run once Firebase credentials are ready.")
        return

    # ---- Real upload path ----
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        sys.exit("Missing dependency. Run: pip install firebase-admin --break-system-packages")

    cred_path = Path(args.credentials)
    if not cred_path.exists():
        sys.exit(
            f"Service account key not found at {cred_path}.\n"
            "Get one from Firebase Console -> Project Settings -> Service Accounts -> "
            "Generate new private key, save it, then re-run with --credentials <path>."
        )

    cred = credentials.Certificate(str(cred_path))
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    for doc in verified_schemes:
        db.collection("schemes").document(doc["schemeId"]).set(doc)
        print(f"  Uploaded scheme: {doc['schemeId']} - {doc['schemeName']}")

    for doc in partner_docs:
        db.collection("channel_partners").document(doc["partnerId"]).set(doc)
        print(f"  Uploaded partner: {doc['partnerId']} - {doc['partnerName']}")

    print(f"\nDone. {len(verified_schemes)} schemes and {len(partner_docs)} channel partners are live in Firestore.")


if __name__ == "__main__":
    main()
